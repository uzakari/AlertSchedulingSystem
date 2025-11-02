#!/usr/bin/env python3
"""
Generate sample Excel file for Alert Scheduling System testing
Requires: pip install openpyxl
"""

import openpyxl
from datetime import datetime
import os

def generate_sample_excel():
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alerts"

    # Headers
    headers = ["AlertType", "Recipient", "EmployeeName", "Department", "DateOfBirth", "StartDate", "YearsOfService", "Title", "Description", "Timestamp", "Location", "StartTime", "EndTime"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Sample data
    sample_data = [
        # System alerts
        ["critical", "+1234567890", "", "", "", "", "", "Database Server Down", "Primary database server is not responding", "2025-01-15 14:30:00", "DataCenter-East", "", ""],
        ["warning", "admin@company.com", "", "", "", "", "", "High CPU Usage", "Server CPU usage has exceeded 85% for 15 minutes", "2025-01-15 14:45:00", "Server-Web01", "", ""],
        ["info", "+1987654321", "", "", "", "", "", "Backup Completed", "Daily backup process completed successfully", "2025-01-15 02:00:00", "Backup-Server", "", ""],
        ["maintenance", "users@company.com", "", "", "", "", "", "Scheduled Maintenance", "System will be offline for routine maintenance", "2025-01-16 08:00:00", "All Systems", "2025-01-16 02:00:00", "2025-01-16 06:00:00"],

        # HR notifications
        ["birthday", "john.doe@company.com", "John Doe", "IT Department", "1985-01-20", "", "", "", "", "2025-01-20 09:00:00", "", "", ""],
        ["birthday", "sarah.smith@company.com", "Sarah Smith", "Marketing", "1990-03-15", "", "", "", "", "2025-03-15 09:00:00", "", "", ""],
        ["anniversary", "mike.johnson@company.com", "Mike Johnson", "Finance", "", "2020-01-18", "5", "", "", "2025-01-18 09:00:00", "", "", ""],
        ["anniversary", "emma.wilson@company.com", "Emma Wilson", "HR", "", "2022-06-01", "3", "", "", "2025-06-01 09:00:00", "", "", ""],

        # Additional system alerts
        ["critical", "emergency@company.com", "", "", "", "", "", "Security Breach", "Unauthorized access attempt detected", "2025-01-15 16:20:00", "Firewall-DMZ", "", ""],
        ["warning", "+1555123456", "", "", "", "", "", "Disk Space Low", "Disk usage on /var partition is at 90%", "2025-01-15 17:00:00", "File-Server", "", ""]
    ]

    # Add data rows
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Create output directory and file
    output_dir = "sample-excel-files"
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"sample_alerts_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    # Save the file
    wb.save(filepath)

    print(f"Sample Excel file generated: {filepath}")
    print("\nFile contains the following alert types:")
    print("- Critical: Database server down, Security breach")
    print("- Warning: High CPU usage, Disk space low")
    print("- Info: Backup completed")
    print("- Maintenance: Scheduled maintenance")
    print("- Birthday: Employee birthday notifications")
    print("- Anniversary: Employee work anniversary notifications")
    print(f"\nTotal alerts: {len(sample_data)}")
    print("\nColumn structure:")
    print("AlertType | Recipient | EmployeeName | Department | DateOfBirth | StartDate | YearsOfService | Title | Description | Timestamp | Location | StartTime | EndTime")

    return filepath

if __name__ == "__main__":
    generate_sample_excel()